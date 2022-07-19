#########################################################
# Matthew Buchanan
# Working with excel files
#########################################################

import os
import openpyxl
import logging


# Function to fetch all .xlsx file paths with "expedia_report_monthly" name in the specified directory into a list
def fetch_files(f_path):
    logging.info('Fetching files for processing')
    # Check if a file path has been specified, if not use the current working directory
    if f_path == "":
        f_path = os.getcwd() + "\\"
    file_list = []
    # Collect a list of all file paths for .xlsx files found in the directory
    try:
        logging.info('Searching directory: ' + f_path)
        file_list = [f_path + f for f in os.listdir(f_path)
                     if f.startswith('expedia_report_monthly') and f.endswith('2018.xlsx')]
    except OSError:
        logging.error('Invalid directory specified, terminating program')
        quit()
    # If no files with .xlsx extension and proper name were found, exit program. Else return the list of file paths
    if len(file_list) == 0:
        logging.error('No files with .xlsx extension found, terminating program')
        quit()
    else:
        for f in file_list:
            logging.info('File found: ' + str(f))
        return file_list


# Function to process each .xlsx file
def process_files(xlsx_file_list):
    logging.info('Preparing files for processing')
    for wb in xlsx_file_list:
        #  Open each file in the file list for processing
        try:
            logging.info('Opening file ' + wb)
            current_wb = openpyxl.load_workbook(filename=wb)
        except openpyxl.utils.exceptions.InvalidFileException:
            logging.error('File' + wb + 'is invalid or corrupted, cannot process')
            break  # This file is no good, move to next file in file list
        #  Prepare to process the sheet titled 'Summary Rolling MoM'
        try:
            current_sheet = current_wb['Summary Rolling MoM']
            logging.info('Sheet titled Summary Rolling MoM found')
            process_sheet_summary_rolling_mom(current_sheet)
        except KeyError:
            logging.warning('No sheet with name Summary Rolling MoM found in the current work book')
        #  Prepare to process the sheet titled 'VOC Rolling MoM'
        try:
            current_sheet = current_wb['VOC Rolling MoM']
            logging.info('Sheet titled VOC Rolling MoM found')
            process_sheet_voc_rolling_mom(current_sheet)
        except KeyError:
            logging.warning('No sheet with name VOC Rolling MoM found in the current work book')


# Function to process the sheet titled 'Summary Rolling MoM'
def process_sheet_summary_rolling_mom(current_sheet):
    # Search each cell in each column using nested loop and break out of both loops once row with '2018-01' is found
    logging.info('Begin processing sheet titled "Summary Rolling MoM" ')
    current_row = -1  # -1 is a sentinel value indicating the desired row was not found
    break_out_var = False  # We use this only to break out of the outer loop once desired row number is found
    for col in current_sheet.iter_cols():
        for cell in col:
            if str(cell.value).startswith('2018-01'):
                logging.info('Found row for January 2018')
                current_row = cell.row
                break_out_var = True
                break
        if break_out_var:
            break  # Break out of the outer loop as well once desired cell is found
    # If the desired row was found log all cells with values in the row along with their headers
    if current_row >= 0:
        values = [str(current_sheet.cell(1, cell.column).value).strip() + ': ' + str(cell.value)
                  for cell in current_sheet[current_row] if cell.value and cell.column != 1]
        for i in range(0, len(values)):
            logging.info(values[i].strip())
        logging.info('Finished processing sheet Summary Rolling MoM')
    else:
        logging.warning('The desired row was not found in the sheet Summary Rolling MoM')


# Function to process the sheet titled 'VOC Rolling MoM'
def process_sheet_voc_rolling_mom(current_sheet):
    # Search each cell using nested loop and break out of both loops once row with '2018-01' or 'January" is found
    logging.info('Begin processing sheet titled "VOC Rolling MoM"')
    current_column = ''  # '' is a sentinel value indicating desired column was not found
    break_out_var = False  # We use this only to break out of the outer loop once desired column number is found
    for row in current_sheet.iter_rows():
        for cell in row:
            if str(cell.value).startswith('2018-01') or str(cell.value) == 'January':
                logging.info('Found column for January 2018')
                current_column = cell.column_letter
                break_out_var = True
                break
        if break_out_var:
            break  # Break out of the outer loop as well once desired cell is found
    # If the desired column was found log all cells with values in the column along with their row labels
    if current_column != '':
        values = [str(current_sheet.cell(cell.row, 1).value).strip() + ': ' + str(cell.value)
                  for cell in current_sheet[current_column] if (cell.value or current_sheet.cell(cell.row, 1).value)
                  and cell.row != 1]
        for i in range(0, len(values)):
            logging.info(values[i].strip())
        logging.info('Finished processing sheet VOC Rolling MoM')
    else:
        logging.warning('The desired column was not found the sheet VOC Rolling MoM')


def main():
    file_path = ""  # Update if a specific file path is needed, else leave blank
    logging.basicConfig(filename='log_file.log', encoding='utf-8',  # Set up the log-file generator
                        format='%(asctime)s %(levelname)-4s %(message)s', level=logging.DEBUG)
    logging.info('Starting program')
    process_files(fetch_files(file_path))
    logging.info("Program Completed Successfully")


if __name__ == "__main__":
    main()
